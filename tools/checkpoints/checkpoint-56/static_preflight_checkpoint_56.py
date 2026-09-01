#!/usr/bin/env python3
from __future__ import annotations
import argparse,csv,hashlib,json,re,sys
from collections import Counter
from pathlib import Path
from docx import Document
from jsonschema import Draft202012Validator,Draft7Validator
from openpyxl import load_workbook
from pygments import lex
from pygments.lexers.dotnet import CSharpLexer
from pygments.token import Comment,Literal,String

parser=argparse.ArgumentParser(); parser.add_argument('--root',default=None); args=parser.parse_args()
ROOT=Path(args.root).resolve() if args.root else Path(__file__).resolve().parents[3]
PT=ROOT/'docs/design/player_technology'; AT=ROOT/'src/StarCluster.ScenarioRunner/Scenarios/ArchitectureTechnology'
OUT=ROOT/'checkpoint-56-static-preflight.txt'
checks=[]
def ok(name,cond,detail=''): checks.append((name,bool(cond),'' if detail=='' else str(detail)))
def loadj(p): return json.loads(p.read_text(encoding='utf-8-sig'))
def read(p): return p.read_text(encoding='utf-8-sig')
def sha(p):
 h=hashlib.sha256()
 with p.open('rb') as f:
  for c in iter(lambda:f.read(1024*1024),b''): h.update(c)
 return h.hexdigest()
def optional(v,k,d=None): return v[k] if k in v else d

def validate_hash_list(path,expected_count,label):
 lines=[x for x in read(path).splitlines() if x.strip()]
 ok(f'{label} count',len(lines)==expected_count,len(lines))
 for line in lines:
  m=re.match(r'^([0-9a-fA-F]{64})  (.+)$',line); ok(f'{label} parse {line[-70:]}',m is not None)
  if not m: continue
  digest,rel=m.group(1).lower(),m.group(2); p=ROOT/rel
  ok(f'{label} exists {rel}',p.is_file())
  if p.is_file(): ok(f'{label} hash {rel}',sha(p)==digest,sha(p))

required=[
 ROOT/'Checkpoint_56_Readme.txt',ROOT/'README.md',ROOT/'docs/README.md',ROOT/'docs/Prototype_TODO.md',
 ROOT/'docs/Star_Cluster_Game_Concept_v0.5c.docx',PT/'README.md',PT/'StarCluster_Player_TL_Framework_Draft_v0_37.xlsx',
 PT/'Player_TL1_TL9_Technology_Architecture_v0_8.md',PT/'player_technology_architecture_v0_8.json',PT/'player_technology_architecture_schema_v0_8.json',
 PT/'scenario_architecture_bridge_v0_8.json',PT/'checkpoint_56_tl3_defensive_microstep_candidates_v0_1.json',PT/'checkpoint_56_power_aux_characteristic_sweep_v0_1.json',PT/'checkpoint_56_equal_capacity_power_loadouts_v0_1.csv',
 ROOT/'docs/validation/Checkpoint_56_TL3_Defensive_Microsteps_And_Independent_Power_AUX_Screening.md',
 AT/'tl1-tl3-standard-runtime-profiles-v0_3.json',AT/'tl2-tl3-production-auxiliary-profiles-v0_1.json',AT/'tl3-power-component-sweep-profiles-v0_1.json',
 AT/'tl3-itc04-defensive-microstep-screening.json',AT/'tl3-aux04-offense-base-two-capacity-screening.json',AT/'tl3-aux05-shield-breakpoint-screening.json',AT/'tl3-aux06-tl2-tl3-production-progression.json',AT/'tl3-pwr03-component-characteristic-sweep.json',AT/'tl3-pwr04-equal-capacity-power-loadouts.json',
 ROOT/'tools/calibration/checkpoints/checkpoint-56.json',ROOT/'tools/checkpoints/checkpoint-56/apply_checkpoint_56.ps1',ROOT/'tools/checkpoints/checkpoint-56/test_technology_architecture.ps1',ROOT/'tools/checkpoints/checkpoint-56/static_preflight_checkpoint_56.py',ROOT/'tools/checkpoints/checkpoint-56/checkpoint_55b_scenario_hashes.txt',ROOT/'tools/checkpoints/checkpoint-56/checkpoint_55b_critical_hashes.txt',
 ROOT/'src/StarCluster.ScenarioRunner/AuxiliaryTechnology/AuxiliaryCombatProfileCatalog.cs',ROOT/'src/StarCluster.ScenarioRunner/TL1Calibration/Tl1IntegratedTacticalCombatRunner.cs',
]
for p in required: ok(f'required file {p.relative_to(ROOT)}',p.is_file())

# Parse all repository JSON.
for p in sorted(ROOT.rglob('*.json')):
 if any(part in {'out','bin','obj','.git'} for part in p.parts): continue
 try: loadj(p); ok(f'JSON parse {p.relative_to(ROOT)}',True)
 except Exception as e: ok(f'JSON parse {p.relative_to(ROOT)}',False,e)

# Frozen accepted boundary and all unrelated source/test code.
validate_hash_list(ROOT/'tools/checkpoints/checkpoint-56/checkpoint_55b_scenario_hashes.txt',79,'frozen 55b scenario')
validate_hash_list(ROOT/'tools/checkpoints/checkpoint-56/checkpoint_55b_critical_hashes.txt',467,'frozen 55b critical')

# Architecture/schema/bridge.
a=loadj(PT/'player_technology_architecture_v0_8.json'); schema=loadj(PT/'player_technology_architecture_schema_v0_8.json')
try:
 Validator=Draft202012Validator if schema.get('$schema','').endswith('2020-12/schema') else Draft7Validator
 errs=sorted(Validator(schema).iter_errors(a),key=lambda e:list(e.path)); ok('architecture schema validation',not errs,'; '.join(e.message for e in errs[:5]))
except Exception as e: ok('architecture schema validation',False,e)
ok('architecture identity',a.get('id')=='player-technology-architecture-v0_8'); ok('architecture checkpoint',a.get('checkpoint')==56); ok('architecture status',a.get('status')=='tl3_defensive_microstep_and_independent_power_aux_screening')
ew=[1,1,1,2,2,2,3,3,3]; ea=[1,1,2,2,2,3,3,3,4]; cap=a['installationCapacityProposals']
ok('weapon capacity curve',[cap['weaponBayCapacity'][str(i)] for i in range(1,10)]==ew,cap['weaponBayCapacity']); ok('AUX capacity curve',[cap['auxiliaryCapacity'][str(i)] for i in range(1,10)]==ea,cap['auxiliaryCapacity'])
review=loadj(PT/'cruiser_installation_capacity_review_v0_2.json'); ok('capacity review weapon curve',[review['capacityCurve']['weaponBayCapacity'][str(i)] for i in range(1,10)]==ew); ok('capacity review AUX curve',[review['capacityCurve']['auxiliaryCapacity'][str(i)] for i in range(1,10)]==ea)
principles=' '.join(a.get('principles',[])); ok('independent AUX principle','Every installed AUX system is an independent system' in principles); ok('equal-capacity principle','equal installation opportunity cost' in principles); ok('duplicate AUX legal principle','Duplicate AUX installations are legal' in principles)
subs={x['id']:x for x in a['subfamilies']}; ok('battery TL3 sweep documented','3/4 charges' in subs['aux_combat_battery']['milestones']['3']); ok('capacitor TL3 sweep documented','capacity 2/3' in subs['aux_power_capacitor']['milestones']['3'])
b=loadj(PT/'scenario_architecture_bridge_v0_8.json'); ok('bridge checkpoint/status',b['checkpoint']==56 and b['status']=='tl1_tl2_frozen_tl3_microstep_and_independent_power_screening_bridge'); ok('bridge standard catalog v0_3',b['standardProfileCatalog'].endswith('tl1-tl3-standard-runtime-profiles-v0_3.json')); ok('bridge TL3 single-main',b['matrixPolicy']['normalTl3WeaponBays']==1); ok('bridge no auto promotion',b['matrixPolicy']['automaticPromotion'] is False)
for k,rel in b.get('companionFiles',{}).items(): ok(f'bridge companion {k}',(ROOT/rel).is_file(),rel)

# Standard profile microsteps.
std=loadj(AT/'tl1-tl3-standard-runtime-profiles-v0_3.json'); sm={p['id']:p for p in std['profiles']}; ok('standard v0_3 count unique',len(std['profiles'])==7 and len(sm)==7,(len(std['profiles']),len(sm)))
for pid in ['tl1-production','tl2-production','tl3-lowtech-control','tl3-offense-refinement','tl3-offense-plus-hull1','tl3-offense-plus-ai1','tl3-offense-plus-shield1']: ok(f'standard profile {pid}',pid in sm)
off=sm['tl3-offense-refinement']; h=sm['tl3-offense-plus-hull1']; ai=sm['tl3-offense-plus-ai1']; sh=sm['tl3-offense-plus-shield1']
def same_except_def_field(x,y,field):
 for sec in ['powerAndControl','movement','weapons']:
  if x[sec]!=y[sec]: return False
 for k,v in x['defense'].items():
  if k==field:
   if y['defense'][k]!=v+1: return False
  elif y['defense'][k]!=v: return False
 return True
ok('Hull microstep exact',same_except_def_field(off,h,'hull')); ok('AI microstep exact',same_except_def_field(off,ai,'armorIntegrity')); ok('Shield microstep exact',same_except_def_field(off,sh,'shieldCapacity'))

# AUX catalogs.
t3aux=loadj(AT/'tl3-auxiliary-capstone-profiles-v0_2.json'); t3m={p['id']:p for p in t3aux['profiles']}; t3cap2=[p for p in t3aux['profiles'] if p['technologyLevel']==3 and not p['counterfactual'] and p['capacityCost']==2]; ok('TL3 retained capacity2 loadouts',len(t3cap2)==13,len(t3cap2))
cross=loadj(AT/'tl2-tl3-production-auxiliary-profiles-v0_1.json'); ok('cross catalog unique',len({p['id'] for p in cross['profiles']})==len(cross['profiles'])); ok('cross TL2 count',len([p for p in cross['profiles'] if p['technologyLevel']==2 and not p['counterfactual']])==9); ok('cross TL3 cap2 count',len([p for p in cross['profiles'] if p['technologyLevel']==3 and not p['counterfactual'] and p['capacityCost']==2])==13)
powc=loadj(AT/'tl3-power-component-sweep-profiles-v0_1.json'); pm={p['id']:p for p in powc['profiles']}; ok('power catalog profile count unique',len(powc['profiles'])==34 and len(pm)==34,(len(powc['profiles']),len(pm)))
atomic=['aux-r56-b3g1','aux-r56-b4g1','aux-r56-b3g2','aux-r56-b4g2','aux-r56-c2d1','aux-r56-c3d1','aux-r56-c2d2','aux-r56-c3d2']
for pid in atomic:
 p=pm[pid]; ok(f'atomic one slot/component {pid}',p['capacityCost']==1 and len(p.get('powerComponents',[]))==1,(p['capacityCost'],len(p.get('powerComponents',[]))))
# exact candidate characteristics
expect={'aux-r56-b3g1':('CombatBattery',3,1),'aux-r56-b4g1':('CombatBattery',4,1),'aux-r56-b3g2':('CombatBattery',3,2),'aux-r56-b4g2':('CombatBattery',4,2)}
for pid,(kind,charges,gain) in expect.items():
 c=pm[pid]['powerComponents'][0]; ok(f'battery candidate {pid}',c['kind']==kind and c['combatBatteryCharges']==charges and c['combatBatteryGain']==gain,c)
expectc={'aux-r56-c2d1':(2,1),'aux-r56-c3d1':(3,1),'aux-r56-c2d2':(2,2),'aux-r56-c3d2':(3,2)}
for pid,(capacity,discharge) in expectc.items():
 c=pm[pid]['powerComponents'][0]; ok(f'capacitor candidate {pid}',c['kind']=='PowerCapacitor' and c['capacitorCapacity']==capacity and c['capacitorDischargeRate']==discharge and c['capacitorChargeRate']==1,c)
cap2=[p for p in powc['profiles'] if not p['counterfactual'] and p['capacityCost']==2]; ok('power equal-cap count',len(cap2)==25,len(cap2)); reactor=pm['aux-r56-auxiliary-reactor']; ok('reactor control',reactor['auxiliaryReactorOutput']==1 and reactor['capacityCost']==2 and not reactor.get('powerComponents'))
two=[p for p in cap2 if p['id']!='aux-r56-auxiliary-reactor']; ok('two-component alternatives',len(two)==24,len(two))
for p in two:
 comps=p.get('powerComponents',[]); ok(f'independent component count {p["id"]}',len(comps)==2,len(comps)); ok(f'independent component IDs {p["id"]}',len({c['id'] for c in comps})==2,[c['id'] for c in comps])
ok('BB count',len([p for p in two if p['id'].startswith('aux-r56-bb-')])==4); ok('CC count',len([p for p in two if p['id'].startswith('aux-r56-cc-')])==4); ok('BC count',len([p for p in two if p['id'].startswith('aux-r56-bc-')])==16)

# Study envelope + variant legality.
runner=read(ROOT/'src/StarCluster.ScenarioRunner/TL1Calibration/Tl1IntegratedTacticalCombatRunner.cs'); m=re.search(r'private const string SchemaVersion\s*=\s*\n?\s*"([^"]+)";',runner); canonical=m.group(1) if m else ''
baseline=sha(PT/'tl1_core_combat_numerical_baseline_v0_1.csv')
studies={
 'tl3-itc04-defensive-microstep-screening.json':('tl3-itc04-defensive-microstep-screening',560100,108,'tl1-tl3-standard-runtime-profiles-v0_3.json','tl3-auxiliary-capstone-profiles-v0_2.json'),
 'tl3-aux04-offense-base-two-capacity-screening.json':('tl3-aux04-offense-base-two-capacity-screening',560200,585,'tl1-tl3-standard-runtime-profiles-v0_3.json','tl3-auxiliary-capstone-profiles-v0_2.json'),
 'tl3-aux05-shield-breakpoint-screening.json':('tl3-aux05-shield-breakpoint-screening',560300,72,'tl1-tl3-standard-runtime-profiles-v0_3.json','tl3-auxiliary-capstone-profiles-v0_2.json'),
 'tl3-aux06-tl2-tl3-production-progression.json':('tl3-aux06-tl2-tl3-production-progression',560400,702,'tl1-tl3-standard-runtime-profiles-v0_3.json','tl2-tl3-production-auxiliary-profiles-v0_1.json'),
 'tl3-pwr03-component-characteristic-sweep.json':('tl3-pwr03-component-characteristic-sweep',560500,168,'tl1-tl3-standard-runtime-profiles-v0_3.json','tl3-power-component-sweep-profiles-v0_1.json'),
 'tl3-pwr04-equal-capacity-power-loadouts.json':('tl3-pwr04-equal-capacity-power-loadouts',560600,360,'tl1-tl3-standard-runtime-profiles-v0_3.json','tl3-power-component-sweep-profiles-v0_1.json'),
}
expected_keys={'schemaVersion','id','checkpoint','baselineSha256','masterSeed','trialsPerVariant','technologyProfileCatalog','auxiliaryProfileCatalog','variants'}
allids=[]
for file,(sid,seed,count,techf,auxf) in studies.items():
 d=loadj(AT/file); ok(f'envelope exact keys {file}',set(d)==expected_keys,sorted(d)); ok(f'envelope schema {file}',d.get('schemaVersion')==canonical,d.get('schemaVersion')); ok(f'envelope id {file}',d.get('id')==sid); ok(f'envelope checkpoint {file}',d.get('checkpoint')==56); ok(f'envelope baseline {file}',d.get('baselineSha256','').lower()==baseline,(d.get('baselineSha256'),baseline)); ok(f'envelope seed {file}',d.get('masterSeed')==seed); ok(f'envelope trials {file}',d.get('trialsPerVariant')==10000); ok(f'envelope tech catalog {file}',d.get('technologyProfileCatalog').endswith(techf),d.get('technologyProfileCatalog')); ok(f'envelope AUX catalog {file}',d.get('auxiliaryProfileCatalog').endswith(auxf),d.get('auxiliaryProfileCatalog')); ok(f'variant count {file}',len(d.get('variants',[]))==count,len(d.get('variants',[])))
 auxmap={p['id']:p for p in loadj(ROOT/d['auxiliaryProfileCatalog'])['profiles']}; techmap={p['id']:p for p in loadj(ROOT/d['technologyProfileCatalog'])['profiles']}
 for v in d['variants']:
  allids.append(v['id']); ok(f'one-main A {v["id"]}',optional(v,'sideASecondaryFamily') is None); ok(f'one-main B {v["id"]}',optional(v,'sideBSecondaryFamily') is None)
  for side in 'AB':
   pid=v[f'side{side}ProfileId']; aid=v[f'side{side}AuxiliaryProfileId']; ok(f'profile exists {v["id"]} {side}',pid in techmap,pid); ok(f'aux exists {v["id"]} {side}',aid in auxmap,aid)
   if pid in techmap and aid in auxmap:
    tl=techmap[pid]['technologyLevel']; limit=2 if tl==3 else 1; ok(f'aux TL legal {v["id"]} {side}',auxmap[aid]['technologyLevel']<=tl,(auxmap[aid]['technologyLevel'],tl)); ok(f'aux cap legal {v["id"]} {side}',auxmap[aid]['capacityCost']<=limit,(auxmap[aid]['capacityCost'],limit))
ok('all CP56 IDs unique',len(allids)==len(set(allids)),len(allids)); ok('CP56 total variants',len(allids)==1995,len(allids))
# exact label shapes
label_expect={
 'tl3-itc04-defensive-microstep-screening.json':{'tl3-r56-micro-vs-tl2':72,'tl3-r56-micro-attribution':18,'tl3-r56-micro-pairwise':18},
 'tl3-aux04-offense-base-two-capacity-screening.json':{'tl3-r56-aux-legal-matrix':507,'tl3-r56-aux-no-aux-diagnostic':78},
 'tl3-aux05-shield-breakpoint-screening.json':{'tl3-r56-shield-vs-none':48,'tl3-r56-booster-vs-stabilizer':24},
 'tl3-aux06-tl2-tl3-production-progression.json':{'tl3-r56-tl2-tl3-production-progression':702},
 'tl3-pwr03-component-characteristic-sweep.json':{'tl3-r56-power-atomic-normal-vs-none':48,'tl3-r56-power-atomic-stress-vs-none':48,'tl3-r56-battery-characteristic-pairwise':36,'tl3-r56-capacitor-characteristic-pairwise':36},
 'tl3-pwr04-equal-capacity-power-loadouts.json':{'tl3-r56-equal-cap-vs-reactor-normal':144,'tl3-r56-equal-cap-vs-reactor-stress':144,'tl3-r56-equal-cap-representative-pairwise':72},
}
for f,e in label_expect.items(): ok(f'label counts {f}',dict(Counter(v['profileLabel'] for v in loadj(AT/f)['variants']))==e,dict(sorted(Counter(v['profileLabel'] for v in loadj(AT/f)['variants']).items())))

# Runner and catalog source contract; lightweight delimiter scan on only modified C# files.
cat=read(ROOT/'src/StarCluster.ScenarioRunner/AuxiliaryTechnology/AuxiliaryCombatProfileCatalog.cs')
for token in ['powerComponents','AuxiliaryPowerComponentKind','CombatBattery','PowerCapacitor','more independent power components than its AUX capacity cost']:
 ok(f'catalog source token {token}',token in cat)
for token in ['Tl3DefensiveMicrostepStudyId','Tl3EqualCapacityPowerStudyId','IndependentCombatBatteryInstallation','IndependentCapacitorInstallation','IndependentPowerComponentFunctional','power-aux-','reusable capacitor energy is spent before finite battery charges']:
 ok(f'runner source token {token}',token in runner)
for p in [ROOT/'src/StarCluster.ScenarioRunner/AuxiliaryTechnology/AuxiliaryCombatProfileCatalog.cs',ROOT/'src/StarCluster.ScenarioRunner/TL1Calibration/Tl1IntegratedTacticalCombatRunner.cs']:
 # Strip comments/strings using pygments and check delimiters in code tokens.
 code=''.join(t for tok,t in lex(read(p),CSharpLexer()) if tok not in Comment and tok not in Literal.String and tok not in String)
 stack=[]; pairs={')':'(',']':'[','}':'{'}; good=True; bad=''
 for i,ch in enumerate(code):
  if ch in '([{': stack.append(ch)
  elif ch in ')]}':
   if not stack or stack[-1]!=pairs[ch]: good=False; bad=f'{ch}@{i}'; break
   stack.pop()
 if stack: good=False; bad=f'unclosed {stack[-10:]}'
 ok(f'C# delimiters {p.name}',good,bad)

# Checkpoint definition accounting.
cp=loadj(ROOT/'tools/calibration/checkpoints/checkpoint-56.json'); sids=[s['id'] for s in cp['stages']]; ok('checkpoint id',cp['checkpointId']=='56'); ok('45 stages',len(sids)==45 and cp['checkpointMetrics']['stageCount']==45,len(sids)); ok('stage IDs unique',len(sids)==len(set(sids))); ok('self-test final',sids[-1]=='runner-self-tests')
trial_variants=sum(int(s.get('metrics',{}).get('variantCount',0)) for s in cp['stages'] if s.get('metrics',{}).get('usesTrials')); ok('checkpoint MC variants',trial_variants==12691 and cp['checkpointMetrics']['monteCarloVariantCount']==12691,trial_variants); ok('checkpoint default trials',cp['checkpointMetrics']['trialsAtDefault']==126910000); ok('checkpoint primary',cp['primaryStudy']=={'id':'tl3-itc04-defensive-microstep-screening','variantCount':108},cp['primaryStudy']); ok('frozen 55b metric',cp['checkpointMetrics']['frozenCheckpoint55bScenarioJsonCount']==79)
expected_tail=['checkpoint-56-tl3-defensive-microsteps','checkpoint-56-tl3-two-capacity-auxiliary-retest','checkpoint-56-tl3-shield-breakpoint-screening','checkpoint-56-tl2-tl3-production-aux-progression','checkpoint-56-tl3-power-characteristic-sweep','checkpoint-56-tl3-equal-capacity-power-loadouts','runner-self-tests']; ok('CP56 tail stages',sids[-7:]==expected_tail,sids[-7:])

# Workbook structural/formula checks (final QA file).
xlsx=PT/'StarCluster_Player_TL_Framework_Draft_v0_37.xlsx'
if xlsx.is_file():
 wbf=load_workbook(xlsx,data_only=False); wbv=load_workbook(xlsx,data_only=True); ok('workbook minimum sheet count',len(wbf.sheetnames)>=77,len(wbf.sheetnames))
 for sname in ['Overview','Design Decisions','Checkpoint 56 Micro Defense','Checkpoint 56 Power Sweep','Checkpoint 56 Equal Power','Checkpoint 56 Cross-TL AUX']: ok(f'workbook sheet {sname}',sname in wbf.sheetnames)
 formulas=[]; missing=[]; errors=[]
 for sn in wbf.sheetnames:
  wf=wbf[sn]; wv=wbv[sn]
  for row in wf.iter_rows():
   for c in row:
    if c.data_type=='f' or (isinstance(c.value,str) and c.value.startswith('=')):
     formulas.append((sn,c.coordinate)); val=wv[c.coordinate].value
     if val is None: missing.append((sn,c.coordinate))
     if isinstance(val,str) and val.startswith('#'): errors.append((sn,c.coordinate,val))
 ok('workbook formula count',len(formulas)==229,len(formulas)); ok('workbook cached formulas complete',not missing,missing[:10]); ok('workbook cached errors absent',not errors,errors[:10])
 ov=' '.join(str(c.value or '') for row in wbf['Overview'].iter_rows() for c in row); ok('workbook CP56 version','v0.37' in str(wbf['Overview']['A1'].value) and 'Checkpoint 56' in ov)
 dd=wbf['Design Decisions']; ids=[str(dd.cell(r,1).value or '') for r in range(1,dd.max_row+1)]
 for n in range(518,528): ok(f'workbook D-{n}',ids.count(f'D-{n}')==1)
else: ok('workbook exists',False)

# Concept structural text checks.
docx=ROOT/'docs/Star_Cluster_Game_Concept_v0.5c.docx'
if docx.is_file():
 doc=Document(docx); text='\n'.join(p.text for p in doc.paragraphs)
 ok('concept v0.5c marker','v0.5c' in text); ok('concept CP56 section','Checkpoint 56' in text and 'Per-installed-AUX independence' in text); ok('concept equal-capacity power','equal-capacity' in text.lower() or 'equal capacity' in text.lower())
 for n in range(518,528): ok(f'concept D-{n}',text.count(f'D-{n}')==1,text.count(f'D-{n}'))
else: ok('concept exists',False)

# Documentation front doors.
for p in [ROOT/'README.md',ROOT/'docs/README.md',ROOT/'docs/Prototype_TODO.md',PT/'README.md',ROOT/'docs/validation/Checkpoint_56_TL3_Defensive_Microsteps_And_Independent_Power_AUX_Screening.md']:
 if p.is_file():
  t=read(p); ok(f'front door CP56 {p.relative_to(ROOT)}','Checkpoint 56' in t or 'checkpoint-56' in t.lower())

passed=sum(1 for _,v,_ in checks if v); failed=[x for x in checks if not x[1]]
lines=[f'Checkpoint 56 static preflight: {passed}/{len(checks)} checks passed; {len(failed)} failed.']
for name,val,detail in checks:
 lines.append(('PASS ' if val else 'FAIL ')+name+(f' :: {detail}' if detail else ''))
OUT.write_text('\n'.join(lines)+'\n',encoding='utf-8')
print(lines[0])
if failed:
 for x in failed[:25]: print('FAIL',x[0],x[2])
 sys.exit(1)
