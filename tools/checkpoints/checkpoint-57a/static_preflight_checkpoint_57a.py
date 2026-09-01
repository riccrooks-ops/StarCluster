#!/usr/bin/env python3
from __future__ import annotations
import argparse,csv,hashlib,json,re,sys
from collections import Counter
from pathlib import Path
from docx import Document
from jsonschema import Draft202012Validator,Draft7Validator
from openpyxl import load_workbook
from pygments import lex
from pygments.lexers.dotnet import CSharpLexer
from pygments.token import Comment,Literal,String

parser=argparse.ArgumentParser(); parser.add_argument('--root',default=None); args=parser.parse_args()
ROOT=Path(args.root).resolve() if args.root else Path(__file__).resolve().parents[3]
PT=ROOT/'docs/design/player_technology'; AT=ROOT/'src/StarCluster.ScenarioRunner/Scenarios/ArchitectureTechnology'
OUT=ROOT/'checkpoint-57a-static-preflight.txt'
checks=[]
def ok(name,cond,detail=''): checks.append((name,bool(cond),'' if detail=='' else str(detail)))
def loadj(p): return json.loads(p.read_text(encoding='utf-8-sig'))
def read(p): return p.read_text(encoding='utf-8-sig')
def sha(p):
 h=hashlib.sha256()
 with p.open('rb') as f:
  for c in iter(lambda:f.read(1024*1024),b''): h.update(c)
 return h.hexdigest()
def opt(v,k,d=None): return v[k] if k in v else d

def validate_hash_list(path,expected_count,label):
 lines=[x for x in read(path).splitlines() if x.strip()] if path.is_file() else []
 ok(f'{label} file',path.is_file()); ok(f'{label} count',len(lines)==expected_count,len(lines))
 for line in lines:
  m=re.match(r'^([0-9a-fA-F]{64})  (.+)$',line); ok(f'{label} parse {line[-72:]}',m is not None)
  if not m: continue
  digest,rel=m.group(1).lower(),m.group(2); p=ROOT/rel
  ok(f'{label} exists {rel}',p.is_file())
  if p.is_file(): ok(f'{label} hash {rel}',sha(p)==digest,sha(p))

required=[
 ROOT/'Checkpoint_57a_Readme.txt',ROOT/'README.md',ROOT/'docs/README.md',ROOT/'docs/Prototype_TODO.md',
 ROOT/'docs/Star_Cluster_Game_Concept_v0.5d.docx',PT/'README.md',PT/'StarCluster_Player_TL_Framework_Draft_v0_38.xlsx',
 PT/'Player_TL1_TL9_Technology_Architecture_v0_9.md',PT/'player_technology_architecture_v0_9.json',PT/'player_technology_architecture_schema_v0_9.json',
 PT/'scenario_architecture_bridge_v0_9.json',PT/'checkpoint_57_tl4_foundation_transition_v0_1.json',PT/'checkpoint_57_tl4_power_loadouts_v0_1.csv',
 ROOT/'docs/validation/Checkpoint_57a_TL4_Dual_Main_Compile_And_Launcher_Hotfix.md',
 AT/'tl1-tl4-standard-runtime-profiles-v0_1.json',AT/'tl3-tl4-production-auxiliary-profiles-v0_1.json',
 AT/'tl4-itc01-foundation-transition.json',AT/'tl4-itc02-two-bay-loadout-screening.json',AT/'tl4-itc03-tl3-specialization-resistance.json',AT/'tl4-pwr01-natural-two-bay-power.json',AT/'tl4-pwr02-mixed-power-flexibility.json',
 ROOT/'tools/calibration/checkpoints/checkpoint-57a.json',ROOT/'tools/checkpoints/checkpoint-57a/apply_checkpoint_57a.ps1',ROOT/'tools/checkpoints/checkpoint-57a/test_technology_architecture.ps1',ROOT/'tools/checkpoints/checkpoint-57a/static_preflight_checkpoint_57a.py',ROOT/'tools/checkpoints/checkpoint-57a/checkpoint_56_scenario_hashes.txt',ROOT/'tools/checkpoints/checkpoint-57a/checkpoint_56_critical_hashes.txt',
 ROOT/'src/StarCluster.ScenarioRunner/TL1Calibration/Tl1IntegratedTacticalCombatRunner.cs',
]
for p in required: ok(f'required file {p.relative_to(ROOT)}',p.is_file())

# Parse all repository JSON to catch malformed additive or historical documents.
for p in sorted(ROOT.rglob('*.json')):
 if any(part in {'out','bin','obj','.git'} for part in p.parts): continue
 try: loadj(p); ok(f'JSON parse {p.relative_to(ROOT)}',True)
 except Exception as e: ok(f'JSON parse {p.relative_to(ROOT)}',False,e)

# Freeze accepted CP56 scenarios and all unrelated source/test/reference files.
validate_hash_list(ROOT/'tools/checkpoints/checkpoint-57a/checkpoint_56_scenario_hashes.txt',88,'frozen 56 scenario')
validate_hash_list(ROOT/'tools/checkpoints/checkpoint-57a/checkpoint_56_critical_hashes.txt',495,'frozen 56 critical')

# Architecture/schema/bridge.
a=loadj(PT/'player_technology_architecture_v0_9.json'); schema=loadj(PT/'player_technology_architecture_schema_v0_9.json')
try:
 Validator=Draft202012Validator if schema.get('$schema','').endswith('2020-12/schema') else Draft7Validator
 errs=sorted(Validator(schema).iter_errors(a),key=lambda e:list(e.path)); ok('architecture schema validation',not errs,'; '.join(e.message for e in errs[:5]))
except Exception as e: ok('architecture schema validation',False,e)
ok('architecture identity',a.get('id')=='player-technology-architecture-v0_9',a.get('id')); ok('architecture checkpoint',a.get('checkpoint')==57,a.get('checkpoint')); ok('architecture status',a.get('status')=='tl1_tl3_frozen_tl4_dual_main_foundation_screening',a.get('status'))
ew=[1,1,1,2,2,2,3,3,3]; ea=[1,1,2,2,2,3,3,3,4]; cap=a['installationCapacityProposals']
ok('weapon capacity curve',[cap['weaponBayCapacity'][str(i)] for i in range(1,10)]==ew,cap['weaponBayCapacity']); ok('AUX capacity curve',[cap['auxiliaryCapacity'][str(i)] for i in range(1,10)]==ea,cap['auxiliaryCapacity'])
principles=' '.join(a.get('principles',[])); ok('TL3 frozen principle','Checkpoint 56 closes TL3' in principles); ok('TL4 dual-main principle','two unrestricted main weapons' in principles); ok('natural power principle','No synthetic background Tactical Power' in principles); ok('single-main fallback principle','one main weapon throughout TL1-TL9' in principles)
subs={x['id']:x for x in a['subfamilies']}; ok('Battery B4G1 frozen','B4G1' in json.dumps(subs.get('aux_combat_battery',{}))); ok('Capacitor C2D1 frozen','C2D1' in json.dumps(subs.get('aux_power_capacitor',{})))
bridge=loadj(PT/'scenario_architecture_bridge_v0_9.json'); ok('bridge checkpoint/status',bridge.get('checkpoint')==57 and bridge.get('status')=='tl1_tl3_frozen_tl4_dual_main_foundation_screening_bridge',(bridge.get('checkpoint'),bridge.get('status'))); ok('bridge TL3 main',bridge['matrixPolicy']['normalTl3WeaponBays']==1); ok('bridge TL4 mains',bridge['matrixPolicy']['normalTl4WeaponBays']==2); ok('bridge TL4 AUX',bridge['matrixPolicy']['normalTl4AuxiliaryCapacity']==2); ok('bridge no synthetic TP',bridge['matrixPolicy']['syntheticBackgroundTacticalPower'] is False); ok('bridge no auto promotion',bridge['matrixPolicy']['automaticPromotion'] is False)
for k,rel in bridge.get('companionFiles',{}).items(): ok(f'bridge companion {k}',(ROOT/rel).is_file(),rel)

# Standard production profiles: TL4 is a pure architecture control.
std=loadj(AT/'tl1-tl4-standard-runtime-profiles-v0_1.json'); sm={p['id']:p for p in std['profiles']}; ok('standard profile IDs unique',len(sm)==len(std['profiles']),len(std['profiles']))
for pid in ['tl1-production','tl2-production','tl3-production','tl4-foundation-control']: ok(f'standard profile {pid}',pid in sm)
tl3=sm['tl3-production']; tl4=sm['tl4-foundation-control']
ok('TL3 level',tl3['technologyLevel']==3); ok('TL4 level',tl4['technologyLevel']==4)
for sec in ['defense','powerAndControl','movement','weapons']: ok(f'TL4 equals TL3 {sec}',tl4[sec]==tl3[sec],(tl3[sec],tl4[sec]))
ok('TL3 frozen defense',tl3['defense']['hull']==12 and tl3['defense']['armorIntegrity']==5 and tl3['defense']['shieldCapacity']==2,tl3['defense'])
ok('TL3 frozen offense',tl3['powerAndControl']['targetingBonus']==13 and tl3['weapons']['kinetic']['accuracyBonus']==24 and tl3['weapons']['energy']['accuracyBonus']==29 and tl3['weapons']['missile']['guidanceChance']==62)

# Production AUX catalog and independent power semantics.
auxd=loadj(AT/'tl3-tl4-production-auxiliary-profiles-v0_1.json'); aux={p['id']:p for p in auxd['profiles']}; ok('AUX profile IDs unique',len(aux)==len(auxd['profiles']),(len(aux),len(auxd['profiles'])))
for pid in ['aux-r57-none-tl3','aux-r57-none-tl4','aux-r57-battery','aux-r57-capacitor','aux-r57-reactor','aux-r57-bb','aux-r57-cc','aux-r57-bc','aux-r57-battery-evasion','aux-r57-capacitor-evasion','aux-r57-battery-amm','aux-r57-capacitor-epds','aux-r57-battery-epds','aux-r57-capacitor-amm']:
 ok(f'AUX profile {pid}',pid in aux)
bat=aux['aux-r57-battery']; capc=aux['aux-r57-capacitor']; reactor=aux['aux-r57-reactor']
ok('Battery one slot',bat['capacityCost']==1 and len(bat.get('powerComponents',[]))==1); bc=bat['powerComponents'][0]; ok('Battery B4G1 exact',bc['kind']=='CombatBattery' and bc['combatBatteryCharges']==4 and bc['combatBatteryGain']==1,bc)
ok('Capacitor one slot',capc['capacityCost']==1 and len(capc.get('powerComponents',[]))==1); cc=capc['powerComponents'][0]; ok('Capacitor C2D1 exact',cc['kind']=='PowerCapacitor' and cc['capacitorCapacity']==2 and cc['capacitorChargeRate']==1 and cc['capacitorDischargeRate']==1,cc)
ok('Reactor two slot +1',reactor['capacityCost']==2 and reactor['auxiliaryReactorOutput']==1 and not reactor.get('powerComponents'),reactor)
for pid,kinds in [('aux-r57-bb',['CombatBattery','CombatBattery']),('aux-r57-cc',['PowerCapacitor','PowerCapacitor']),('aux-r57-bc',['CombatBattery','PowerCapacitor'])]:
 p=aux[pid]; pcs=p.get('powerComponents',[]); ok(f'{pid} two-slot/two-component',p['capacityCost']==2 and len(pcs)==2,(p['capacityCost'],len(pcs))); ok(f'{pid} independent IDs',len({x['id'] for x in pcs})==2,[x['id'] for x in pcs]); ok(f'{pid} component kinds',[x['kind'] for x in pcs]==kinds,[x['kind'] for x in pcs])
# Mixed flexibility lane deliberately uses only support systems with established independent component representation.
mixed_expected={'aux-r57-battery-evasion','aux-r57-capacitor-evasion','aux-r57-battery-amm','aux-r57-capacitor-epds','aux-r57-battery-epds','aux-r57-capacitor-amm'}
for pid in sorted(mixed_expected):
 p=aux[pid]; ok(f'mixed two-slot {pid}',p['capacityCost']==2); ok(f'mixed one independent power component {pid}',len(p.get('powerComponents',[]))==1,p.get('powerComponents'))

# Canonical integrated-combat study envelopes and per-variant legality.
baseline=sha(PT/'tl1_core_combat_numerical_baseline_v0_1.csv'); canonical='star-cluster-tl1-integrated-tactical-combat-v2'; expected_keys={'schemaVersion','id','checkpoint','baselineSha256','masterSeed','trialsPerVariant','technologyProfileCatalog','auxiliaryProfileCatalog','variants'}
studies={
 'tl4-itc01-foundation-transition.json':('tl4-itc01-foundation-transition',570100,180),
 'tl4-itc02-two-bay-loadout-screening.json':('tl4-itc02-two-bay-loadout-screening',570200,243),
 'tl4-itc03-tl3-specialization-resistance.json':('tl4-itc03-tl3-specialization-resistance',570300,468),
 'tl4-pwr01-natural-two-bay-power.json':('tl4-pwr01-natural-two-bay-power',570400,120),
 'tl4-pwr02-mixed-power-flexibility.json':('tl4-pwr02-mixed-power-flexibility',570500,144),
}
allids=[]
for file,(sid,seed,count) in studies.items():
 d=loadj(AT/file); ok(f'envelope exact keys {file}',set(d)==expected_keys,sorted(d)); ok(f'envelope schema {file}',d.get('schemaVersion')==canonical,d.get('schemaVersion')); ok(f'envelope id {file}',d.get('id')==sid); ok(f'envelope checkpoint {file}',d.get('checkpoint')==57); ok(f'envelope baseline {file}',d.get('baselineSha256','').lower()==baseline,(d.get('baselineSha256'),baseline)); ok(f'envelope seed {file}',d.get('masterSeed')==seed); ok(f'envelope trials {file}',d.get('trialsPerVariant')==10000); ok(f'envelope tech catalog {file}',d.get('technologyProfileCatalog','').endswith('tl1-tl4-standard-runtime-profiles-v0_1.json')); ok(f'envelope AUX catalog {file}',d.get('auxiliaryProfileCatalog','').endswith('tl3-tl4-production-auxiliary-profiles-v0_1.json')); ok(f'variant count {file}',len(d.get('variants',[]))==count,len(d.get('variants',[])))
 techmap={p['id']:p for p in loadj(ROOT/d['technologyProfileCatalog'])['profiles']}; auxmap={p['id']:p for p in loadj(ROOT/d['auxiliaryProfileCatalog'])['profiles']}
 for v in d['variants']:
  vid=v['id']; allids.append(vid); ok(f'variant ID nonempty {vid}',bool(vid)); ok(f'comparison group {vid}',bool(v.get('comparisonGroup'))); ok(f'profile label {vid}',bool(v.get('profileLabel')))
  ok(f'no synthetic TP A {vid}',v.get('sideABackgroundTacticalPowerCommitment',0)==0,v.get('sideABackgroundTacticalPowerCommitment')); ok(f'no synthetic TP B {vid}',v.get('sideBBackgroundTacticalPowerCommitment',0)==0,v.get('sideBBackgroundTacticalPowerCommitment'))
  for side in 'AB':
   pid=v[f'side{side}ProfileId']; aid=v[f'side{side}AuxiliaryProfileId']; ok(f'profile exists {vid} {side}',pid in techmap,pid); ok(f'aux exists {vid} {side}',aid in auxmap,aid)
   if pid in techmap and aid in auxmap:
    tl=techmap[pid]['technologyLevel']; limit=2 if tl in (3,4) else (1 if tl in (1,2) else 0); ok(f'aux TL legal {vid} {side}',auxmap[aid]['technologyLevel']<=tl,(auxmap[aid]['technologyLevel'],tl)); ok(f'aux capacity legal {vid} {side}',auxmap[aid]['capacityCost']<=limit,(auxmap[aid]['capacityCost'],limit))
   sec=opt(v,f'side{side}SecondaryFamily')
   if sec is not None: ok(f'second main only TL4 {vid} {side}',pid=='tl4-foundation-control',pid)
ok('all CP57 IDs unique',len(allids)==len(set(allids)),len(allids)); ok('CP57 total variants',len(allids)==1155,len(allids))

# Exact study shapes / attribution contracts.
shape={
 'tl4-itc01-foundation-transition.json':{'tl4-r57-capacity-control':18,'tl4-r57-dual-main-transition':162},
 'tl4-itc02-two-bay-loadout-screening.json':{'tl4-r57-two-bay-same-tl':243},
 'tl4-itc03-tl3-specialization-resistance.json':{'tl4-r57-tl3-specialization-resistance':468},
 'tl4-pwr01-natural-two-bay-power.json':{'tl4-r57-power-vs-none':48,'tl4-r57-power-pairwise':72},
 'tl4-pwr02-mixed-power-flexibility.json':{'tl4-r57-mixed-vs-reactor':72,'tl4-r57-mixed-vs-none':72},
}
for f,e in shape.items(): ok(f'label counts {f}',dict(Counter(v['profileLabel'] for v in loadj(AT/f)['variants']))==e,dict(Counter(v['profileLabel'] for v in loadj(AT/f)['variants'])))
trans=loadj(AT/'tl4-itc01-foundation-transition.json')['variants']; control=[v for v in trans if v['profileLabel']=='tl4-r57-capacity-control']; dual=[v for v in trans if v['profileLabel']=='tl4-r57-dual-main-transition']; ok('transition 18 controls',len(control)==18); ok('control is single-main',all(opt(v,'sideASecondaryFamily') is None and opt(v,'sideBSecondaryFamily') is None for v in control)); ok('transition 162 dual',len(dual)==162); ok('dual exactly one TL4 side with second main',all(((v['sideAProfileId']=='tl4-foundation-control' and opt(v,'sideASecondaryFamily') is not None and opt(v,'sideBSecondaryFamily') is None) or (v['sideBProfileId']=='tl4-foundation-control' and opt(v,'sideBSecondaryFamily') is not None and opt(v,'sideASecondaryFamily') is None)) for v in dual))
load=loadj(AT/'tl4-itc02-two-bay-loadout-screening.json')['variants']; ordered={(v['sideAFamily'],v['sideASecondaryFamily']) for v in load}; ok('nine ordered two-bay loadouts',len(ordered)==9,sorted(ordered)); ok('two-bay ranges 3/4/5',set(v['initialRangeHexes'] for v in load)=={3,4,5})
res=loadj(AT/'tl4-itc03-tl3-specialization-resistance.json')['variants']; tl3aux={v['sideBAuxiliaryProfileId'] if v['sideBProfileId']=='tl3-production' else v['sideAAuxiliaryProfileId'] for v in res}; ok('resistance 13 TL3 specializations',len(tl3aux)==13,sorted(tl3aux)); ok('resistance TL4 naked',all((v['sideAAuxiliaryProfileId']=='aux-r57-none-tl4' if v['sideAProfileId']=='tl4-foundation-control' else v['sideBAuxiliaryProfileId']=='aux-r57-none-tl4') for v in res))
pwr=loadj(AT/'tl4-pwr01-natural-two-bay-power.json')['variants']; pwrids={v['sideAAuxiliaryProfileId'] for v in pwr}|{v['sideBAuxiliaryProfileId'] for v in pwr}; ok('natural power full builds',{'aux-r57-none-tl4','aux-r57-reactor','aux-r57-bb','aux-r57-cc','aux-r57-bc'}==pwrids,sorted(pwrids)); ok('natural power range4 only',all(v['initialRangeHexes']==4 for v in pwr))
mix=loadj(AT/'tl4-pwr02-mixed-power-flexibility.json')['variants']; mixids=({v['sideAAuxiliaryProfileId'] for v in mix}|{v['sideBAuxiliaryProfileId'] for v in mix})-{'aux-r57-reactor','aux-r57-none-tl4'}; ok('mixed six profiles exact',mixids==mixed_expected,sorted(mixids)); ok('mixed excludes shield-support composite',not ({'aux-r57-battery-shield-booster','aux-r57-capacitor-shield-stabilizer'} & mixids),sorted(mixids)); ok('mixed range4 only',all(v['initialRangeHexes']==4 for v in mix))

# Runner source contract and lightweight syntax scan.
runner_path=ROOT/'src/StarCluster.ScenarioRunner/TL1Calibration/Tl1IntegratedTacticalCombatRunner.cs'; runner=read(runner_path)
for token in ['Tl4FoundationTransitionStudyId','Tl4TwoBayLoadoutStudyId','Tl4SpecializationResistanceStudyId','Tl4NaturalPowerStudyId','Tl4MixedPowerStudyId','ValidateCheckpoint57Coverage','WriteCheckpoint57PowerConfigurations','WriteCheckpoint57Review','IsCheckpoint57Study']:
 ok(f'runner source token {token}',token in runner)
for p in [runner_path]:
 code=''.join(t for tok,t in lex(read(p),CSharpLexer()) if tok not in Comment and tok not in Literal.String and tok not in String)
 stack=[]; pairs={')':'(',']':'[','}':'{'}; good=True; bad=''
 for i,ch in enumerate(code):
  if ch in '([{': stack.append(ch)
  elif ch in ')]}':
   if not stack or stack[-1]!=pairs[ch]: good=False; bad=f'{ch}@{i}'; break
   stack.pop()
 if stack: good=False; bad=f'unclosed {stack[-10:]}'
 ok(f'C# delimiters {p.name}',good,bad)

# Checkpoint 57a compile/launcher hotfix contract.
validate_start=runner.find('    private static void Validate('); validate_end=runner.find('    private static void ValidateTl2CandidateCoverage(',validate_start)
build_start=runner.find('    private static IReadOnlyList<Tl1IntegratedTacticalCombatGate> BuildGates('); build_end=runner.find('    private static ',build_start+20)
ok('locate Validate method',validate_start>=0 and validate_end>validate_start,(validate_start,validate_end)); ok('locate BuildGates method',build_start>=0 and build_end>build_start,(build_start,build_end))
if validate_start>=0 and validate_end>validate_start:
 vb=runner[validate_start:validate_end]; ok('telemetry gate absent from Validate','tl4-foundation-two-main-telemetry' not in vb)
if build_start>=0 and build_end>build_start:
 bb=runner[build_start:build_end]; ok('telemetry gate present in BuildGates','tl4-foundation-two-main-telemetry' in bb)
ok('nullable range format fixed','v.InitialRangeHexes.HasValue ? v.InitialRangeHexes.Value.ToString(CultureInfo.InvariantCulture) : string.Empty' in runner)
ok('invalid nullable ToString absent','v.InitialRangeHexes.ToString(CultureInfo.InvariantCulture)' not in runner)
apply57a=read(ROOT/'tools/checkpoints/checkpoint-57a/apply_checkpoint_57a.ps1')
ok('launcher has no Python dependency',re.search(r'(?im)^\s*&\s*(python|python3|py)(\.exe)?\b',apply57a) is None)
ok('launcher uses 57a definition','checkpoint-57a.json' in apply57a)

# Checkpoint accounting.
cp=loadj(ROOT/'tools/calibration/checkpoints/checkpoint-57a.json'); sids=[s['id'] for s in cp['stages']]; ok('checkpoint id',cp.get('checkpointId')=='57a'); ok('50 stages',len(sids)==50 and cp['checkpointMetrics']['stageCount']==50,len(sids)); ok('stage IDs unique',len(sids)==len(set(sids))); ok('self-test final',sids[-1]=='runner-self-tests',sids[-1])
trial_variants=sum(int(s.get('metrics',{}).get('variantCount',0)) for s in cp['stages'] if s.get('metrics',{}).get('usesTrials')); ok('checkpoint MC variants',trial_variants==13846 and cp['checkpointMetrics']['monteCarloVariantCount']==13846,trial_variants); ok('checkpoint default trials',cp['checkpointMetrics']['trialsAtDefault']==138460000); ok('checkpoint added variants',cp['checkpointMetrics']['checkpoint57AddedMonteCarloVariantCount']==1155); ok('checkpoint frozen CP56 count',cp['checkpointMetrics']['frozenCheckpoint56ScenarioJsonCount']==88); ok('checkpoint primary',cp['primaryStudy']=={'id':'tl4-itc01-foundation-transition','variantCount':180},cp['primaryStudy'])
expected_tail=['checkpoint-57-tl4-foundation-transition','checkpoint-57-tl4-two-bay-loadout-screening','checkpoint-57-tl3-specialization-resistance','checkpoint-57-tl4-natural-power','checkpoint-57-tl4-mixed-power-flexibility','runner-self-tests']; ok('CP57 tail stages',sids[-6:]==expected_tail,sids[-6:])

# Workbook structure and cached formula results.
xlsx=PT/'StarCluster_Player_TL_Framework_Draft_v0_38.xlsx'
if xlsx.is_file():
 wbf=load_workbook(xlsx,data_only=False); wbv=load_workbook(xlsx,data_only=True); ok('workbook exact sheet count',len(wbf.sheetnames)==81,len(wbf.sheetnames))
 for sname in ['Overview','Design Decisions','Checkpoint 57 TL3 Freeze','Checkpoint 57 TL4 Foundation','Checkpoint 57 Natural Power','Checkpoint 57 Study Matrix']: ok(f'workbook sheet {sname}',sname in wbf.sheetnames)
 formulas=[]; missing=[]; errors=[]
 for sn in wbf.sheetnames:
  wf=wbf[sn]; wv=wbv[sn]
  for row in wf.iter_rows():
   for c in row:
    if c.data_type=='f' or (isinstance(c.value,str) and c.value.startswith('=')):
     formulas.append((sn,c.coordinate)); val=wv[c.coordinate].value
     if val is None: missing.append((sn,c.coordinate))
     if isinstance(val,str) and val.startswith('#'): errors.append((sn,c.coordinate,val))
 ok('workbook formula count',len(formulas)==229,len(formulas)); ok('workbook cached formulas complete',not missing,missing[:10]); ok('workbook cached errors absent',not errors,errors[:10]); ov=' '.join(str(c.value or '') for row in wbf['Overview'].iter_rows() for c in row); ok('workbook CP57 version','v0.38' in str(wbf['Overview']['A1'].value) and 'Checkpoint 57' in ov)
 ids=[str(wbf['Design Decisions'].cell(r,1).value or '') for r in range(1,wbf['Design Decisions'].max_row+1)]
 for n in range(528,540): ok(f'workbook D-{n}',ids.count(f'D-{n}')==1,ids.count(f'D-{n}'))
else: ok('workbook exists',False)

# Concept text structure.
docx=ROOT/'docs/Star_Cluster_Game_Concept_v0.5d.docx'
if docx.is_file():
 doc=Document(docx); text='\n'.join(p.text for p in doc.paragraphs); ok('concept v0.5d marker','v0.5d' in text); ok('concept CP57 section','Checkpoint 57' in text and 'TL4' in text and 'two unrestricted main weapons' in text); ok('concept natural power','synthetic background Tactical Power' in text); ok('concept fallback','one main weapon throughout TL1-TL9' in text)
 for n in range(528,540): ok(f'concept D-{n}',text.count(f'D-{n}')==1,text.count(f'D-{n}'))
else: ok('concept exists',False)

# Front-door docs must identify this checkpoint.
for p in [ROOT/'README.md',ROOT/'docs/README.md',ROOT/'docs/Prototype_TODO.md',PT/'README.md',ROOT/'docs/validation/Checkpoint_57a_TL4_Dual_Main_Compile_And_Launcher_Hotfix.md',ROOT/'Checkpoint_57a_Readme.txt']:
 if p.is_file(): ok(f'front door CP57a {p.relative_to(ROOT)}','Checkpoint 57a' in read(p) or 'checkpoint-57a' in read(p).lower())

passed=sum(1 for _,v,_ in checks if v); failed=[x for x in checks if not x[1]]
lines=[f'Checkpoint 57a static preflight: {passed}/{len(checks)} checks passed; {len(failed)} failed.']
for name,val,detail in checks: lines.append(('PASS ' if val else 'FAIL ')+name+(f' :: {detail}' if detail else ''))
OUT.write_text('\n'.join(lines)+'\n',encoding='utf-8')
print(lines[0])
if failed:
 for x in failed[:30]: print('FAIL',x[0],x[2])
 sys.exit(1)
